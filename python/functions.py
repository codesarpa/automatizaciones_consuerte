from config import ENVIRONMENT,URL_SUPERFLEX,URL_HOME,USER,PASS
from playwright.sync_api import Page, expect
# from playwright import playwright 
from playwright.sync_api import sync_playwright
import re
import logging
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import time

playwright = sync_playwright().start()
browser = playwright.chromium.launch(headless=False, args=["--start-maximized"])
page = browser.new_page()

# Rutas de los archivos
ARCHIVO_PROCESADAS = "cedulas_procesadas.xlsx"
ARCHIVO_NOVEDADES = "cedulas_novedades.xlsx"

def registrar_en_excel(archivo, datos, encabezados):
    if not os.path.exists(archivo):
        wb = Workbook()
        ws = wb.active
        ws.append(encabezados)
    else:
        wb = load_workbook(archivo)
        ws = wb.active
    
    ws.append(datos)
    wb.save(archivo)
    wb.close()


def iniciar_sesion_superflex():
    logging.info("Se procede a iniciar sesion")
    try:
        #PRODUCCION
        page.goto(URL_SUPERFLEX)
        
        page.wait_for_timeout(10000)  # Esperar 10 segundos para que la página cargue completamente
        #USUARIO 
        page.wait_for_selector('#float-input',state='visible')
        page.fill('#float-input', USER)
        
        #CONTRASEÑA
        page.wait_for_selector('#float-input-password',state='visible')
        page.fill('#float-input-password', PASS)

        page.wait_for_timeout(10000)  # Esperar 1 segundo para que los campos se llenen
        page.get_by_text("Ingresar").click()

        page.wait_for_url(URL_HOME, timeout=25000)  # Ajusta 
        logging.info("Se incia sesion correctamente")
        print("Login exitoso")
        input("asd")
    except Exception as e:
        logging.error(f"Ocurrio un error al iniciar sesión es superflex: {e}")

def ingresar_menu_relacion_impresora():
    try:
        page.wait_for_timeout(10000)
        RUTA_MENU = '/sf-admin-web-comunes/admin/maestra-equipo-impresora'
        URL_RELACION_EQUIPO_IMPRESORA = f"{URL_SUPERFLEX}{RUTA_MENU}"

        page.goto(URL_RELACION_EQUIPO_IMPRESORA)
        # page.wait_for_timeout(10000)  # Esperar 5 segundos para que la pÃ¡gina cargue completamente
        page.wait_for_url(URL_RELACION_EQUIPO_IMPRESORA, timeout=30000)  # Ajusta
        print("Menu Activar Servicios")
        logging.info("Se ingresa al modulo de Relacion equipo Impresora")
    except Exception as e:
        print(f"â�Œ Error al ingresar al menu de relacion equipo impresora: {e}")
        logging.error(f"Error al ingresar al menu de relacion equipo impresora : {e}")

def crear_configuracion_chance():
    PUERTO_IMPRESION = "LPT"
    NOMBRE_INSTALACION = "PRINTER"
    MODELO_IMPRESORA_ID = "POS_CHANCE_CHANCE"
    TIPO_IMPRESORA_ID = "NUEVO TIPO IMPRESORA PDAS"
    EQUIPO_ID = "40379019"

    time.sleep(2)
    page.get_by_role("button", name="Crear").click()
    time.sleep(2)

    #PUERTO IMPRESION
    input_puerto_impresion = page.get_by_placeholder("Escribir puerto impresion")
    input_puerto_impresion.wait_for(state='visible')
    input_puerto_impresion.fill(PUERTO_IMPRESION)
    time.sleep(2)

    #NOMBRE INSTALACION
    page.get_by_placeholder("Escribir nombre instalacion").fill(NOMBRE_INSTALACION)
    time.sleep(2)

    botones = page.locator("button:has(span.pi.pi-search)")
    inputs = page.locator("input.p-inputtext")
    print(inputs.count())  # Verifica que realmente sean 3

    #MODELO IMPRESOTA ID
    botones.nth(0).click()
    time.sleep(2)
    page.get_by_role("row", name=MODELO_IMPRESORA_ID).locator(".p-radiobutton-box").click()
    time.sleep(2)

    #TIPO IMPRESORA ID
    botones.nth(1).click()
    time.sleep(2)
    page.get_by_role("row", name=TIPO_IMPRESORA_ID).locator(".p-radiobutton-box").click()
    time.sleep(2)

    # #EQUIPO ID
    botones.nth(2).click()
    # input("ENTER PARA CERRAR NAVEGADOR")
    # inputs = page.locator("input.p-inputtext")
    input_equipo_id = inputs.nth(7)
    input_equipo_id.click()
    input_equipo_id.fill(EQUIPO_ID)
    input_equipo_id.press("Enter")
    page.locator("p-tableradiobutton .p-radiobutton-box").click()

    #CLICK EN BOTÓN GUARDAR
    page.get_by_role("button", name="Guardar").click()

    input("ENTER PARA CERRAR NAVEGADOR")
    #CLICK EN BOTÓN CONFIRMAR
    page.locator("button.p-confirm-dialog-accept").click()




    input("ENTER PARA CERRAR NAVEGADOR")

def configuracion_papel_blanco():
    NOMBRE_INSTALACION = "PRINTER1"
    MODELO_IMPRESORA_ID = "IMPRESORA PARA POS PAPEL BLANCO"
    EQUIPO_ID = "40379019"

    #PUERTO IMPRESION
    input_puerto_impresion = page.get_by_placeholder("Escribir puerto impresion")
    input_puerto_impresion.wait_for(state='visible')
    input_puerto_impresion.fill(PUERTO_IMPRESION)

    #NOMBRE INSTALACION
    page.get_by_placeholder("Escribir nombre instalacion").fill(NOMBRE_INSTALACION)
    time.sleep(2)

    botones = page.locator("button:has(span.pi.pi-search)")
    inputs = page.locator("input.p-inputtext")
    print(inputs.count())  # Verifica que realmente sean 3

    #MODELO IMPRESOTA ID
    botones.nth(0).click()
    time.sleep(2)
    page.get_by_role("row", name=MODELO_IMPRESORA_ID).locator(".p-radiobutton-box").click()
    time.sleep(2)

    #TIPO IMPRESORA ID
    botones.nth(1).click()
    time.sleep(2)
    page.get_by_role("row", name=TIPO_IMPRESORA_ID).locator(".p-radiobutton-box").click()
    time.sleep(2)

def configuracion_chance():
    
    #PUERTO IMPRESION
    input_puerto_impresion = page.get_by_placeholder("Escribir puerto impresion")
    input_puerto_impresion.wait_for(state='visible')
    input_puerto_impresion.fill(PUERTO_IMPRESION)

    #NOMBRE INSTALACION
    page.get_by_placeholder("Escribir nombre instalacion").fill(NOMBRE_INSTALACION)
    time.sleep(2)

    botones = page.locator("button:has(span.pi.pi-search)")
    inputs = page.locator("input.p-inputtext")
    print(inputs.count())  # Verifica que realmente sean 3

    #MODELO IMPRESOTA ID
    botones.nth(0).click()
    time.sleep(5)
    page.get_by_role("row", name=MODELO_IMPRESORA_ID).locator(".p-radiobutton-box").click()
    time.sleep(5)

    #TIPO IMPRESORA ID
    botones.nth(1).click()
    time.sleep(5)
    page.get_by_role("row", name=TIPO_IMPRESORA_ID).locator(".p-radiobutton-box").click()
    time.sleep(5)

    #CLICK EN BOTÓN GUARDAR
    # page.get_by_role("button", name="Guardar").click()
#   #EQUIPO ID
    botones.nth(2).click()
    try:
        input_equipo_id = inputs.nth(7)
        input_equipo_id.click()
        input_equipo_id.fill(EQUIPO_ID)
        input_equipo_id.press("Enter")
        page.locator("p-tableradiobutton .p-radiobutton-box").click()
    except:
        return False
    
    valor_input_equipo_id = page.locator('input[placeholder="Buscar equipo id"]').input_value()
    print(valor_input_equipo_id)
    #CLICK EN BOTÓN GUARDAR
    page.get_by_role("button", name="Guardar").click()

    #CLICK EN BOTÓN CONFIRMAR
    page.locator("button.p-confirm-dialog-accept").click()

    # page.get_by_role("button", name="Limpiar").click()
    return valor_input_equipo_id

def crear_configuracion_chance():

    time.sleep(2)
    page.get_by_role("button", name="Crear").click()
    time.sleep(2)

    #PUERTO IMPRESION
    input_puerto_impresion = page.get_by_placeholder("Escribir puerto impresion")
    input_puerto_impresion.wait_for(state='visible')
    input_puerto_impresion.fill(PUERTO_IMPRESION)
    time.sleep(2)

    #NOMBRE INSTALACION
    page.get_by_placeholder("Escribir nombre instalacion").fill(NOMBRE_INSTALACION)
    time.sleep(2)

    botones = page.locator("button:has(span.pi.pi-search)")
    inputs = page.locator("input.p-inputtext")
    print(inputs.count())  # Verifica que realmente sean 3

    #MODELO IMPRESOTA ID
    botones.nth(0).click()
    time.sleep(2)
    page.get_by_role("row", name=MODELO_IMPRESORA_ID).locator(".p-radiobutton-box").click()
    time.sleep(2)

    #TIPO IMPRESORA ID
    botones.nth(1).click()
    time.sleep(2)
    page.get_by_role("row", name=TIPO_IMPRESORA_ID).locator(".p-radiobutton-box").click()
    time.sleep(2)

    # #EQUIPO ID
    botones.nth(2).click()
    # input("ENTER PARA CERRAR NAVEGADOR")
    # inputs = page.locator("input.p-inputtext")
    input_equipo_id = inputs.nth(7)
    input_equipo_id.click()
    input_equipo_id.fill(EQUIPO_ID)
    input_equipo_id.press("Enter")
    page.locator("p-tableradiobutton .p-radiobutton-box").click()

    #CLICK EN BOTÓN GUARDAR
    # page.get_by_role("button", name="Guardar").click()

    input("ENTER PARA CERRAR NAVEGADOR")
    #CLICK EN BOTÓN CONFIRMAR
    # page.locator("button.p-confirm-dialog-accept").click()




    input("ENTER PARA CERRAR NAVEGADOR")
    page.get_by_role("button", name="Limpiar").click()


def prueba():
    EQUIPO_ID = "40379019"

    page.goto('https://prod-superflex-admin.codesa.com.co/sf-admin-web-comunes/admin/maestra-equipo-impresora')
    time.sleep(5)
    botones = page.locator("button:has(span.pi.pi-search)")

    botones.nth(2).click()
    input_equipo_id = page.locator("input.p-inputtext").nth(2)
    input_equipo_id.fill(EQUIPO_ID)
    input_equipo_id.press('Enter')


    input("ENTER PARA CERRAR NAVEGADOR")

# iniciar_sesion_superflex()
# ingresar_menu_relacion_impresora()
# # crear_configuracion_chance()

# df = pd.read_excel('C:/automatizaciones_consuerte/python/superflex/config_impresion_cs10/tat_cs10.xlsx')
# cedulas = df['CÉDULA'].tolist()

# PUERTO_IMPRESION = "LPT"
# TIPO_IMPRESORA_ID = "NUEVO TIPO IMPRESORA PDAS"
# CEDULAS_NOVEDADES = []
# CEDULAS_PROCESADAS = {}
# for cedula in cedulas:
#     cedula = cedula.replace("CV","")
#     EQUIPO_ID = cedula
    
#     for intento in range (1,3):
#         if intento == 1:
#             #CHANCE
#             NOMBRE_INSTALACION = "PRINTER"
#             MODELO_IMPRESORA_ID = "POS_CHANCE_CHANCE"
#         else:
#             #PAPEL BLANCO
#             NOMBRE_INSTALACION = "PRINTER1"
#             MODELO_IMPRESORA_ID = "IMPRESORA PARA POS PAPEL BLANCO"
#         print(intento)
#         print(EQUIPO_ID)
#         print(PUERTO_IMPRESION)
#         print(TIPO_IMPRESORA_ID)
#         print(NOMBRE_INSTALACION)
#         print(MODELO_IMPRESORA_ID)

#         time.sleep(2)
#         page.get_by_role("button", name="Crear").click()
#         time.sleep(2)
#         input_equipo_id = configuracion_chance()
#         if not input_equipo_id:
#             print(f"Error con {cedula}, pasando a la siguiente.")
#             CEDULAS_NOVEDADES.append(cedula)
#             ingresar_menu_relacion_impresora()
#             # Registrar de inmediato en el Excel de novedades
#             registrar_en_excel(
#                 ARCHIVO_NOVEDADES,
#                 [cedula, "Error al configurar"],
#                 ["CÉDULA", "OBSERVACIÓN"]
#             )
#             break
#         else:
#             CEDULAS_PROCESADAS[cedula] = input_equipo_id
#             print(f"Cédula {cedula} procesada con ID {input_equipo_id}")
#             # Registrar de inmediato en el Excel de procesadas
#             registrar_en_excel(
#                 ARCHIVO_PROCESADAS,
#                 [cedula, input_equipo_id],
#                 ["CÉDULA", "ID EQUIPO"]
#             )

# print(CEDULAS_NOVEDADES)
# print(CEDULAS_PROCESADAS)